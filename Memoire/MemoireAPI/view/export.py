import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment
from PIL import Image as PILImage
import io
from django.http import HttpResponse
from django.conf import settings
import os
from base.models import Slambooks, Url_slambook, Question_Options, Questions, Responses, Response_answer
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.views import Response

@api_view(['GET'])
def export_slambook_responses(request, slamid):
    try:
        # Fetch the slambook
        slambook = Url_slambook.objects.get(slamid=slamid).slamid
        slambook_title = slambook.slamtitle

        # Fetch all responses for the slambook
        responses = Responses.objects.filter(slamid=slambook).prefetch_related('response_answer_set__answer_option')

        # Fetch all questions for the slambook
        questions = Questions.objects.filter(slamid=slambook).prefetch_related('options')

        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Slambook Responses"

        # Define headers
        headers = ["Response ID", "Created At"] + [q.questiontext for q in questions]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)  # Make headers bold
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center headers

        # Initialize dictionaries to track maximum widths and heights for each column and row
        max_column_widths = {col: len(str(header)) for col, header in enumerate(headers, 1)}
        max_row_heights = {1: 20}  # Default height for header row

        # Populate data and calculate sizes
        row = 2
        for response in responses:
            # Default row height for non-image content
            max_row_heights[row] = 20  # Default height for text rows

            # Basic response info
            ws.cell(row=row, column=1).value = str(response.responseid)
            ws.cell(row=row, column=2).value = response.created.strftime('%Y-%m-%d %H:%M:%S')

            # Update max widths for Response ID and Created At
            max_column_widths[1] = max(max_column_widths[1], len(str(response.responseid)))
            max_column_widths[2] = max(max_column_widths[2], len(response.created.strftime('%Y-%m-%d %H:%M:%S')))

            # Fetch answers for this response
            answers = Response_answer.objects.filter(responseid=response)

            # Map answers to questions
            for col, question in enumerate(questions, 3):
                answer = next((a for a in answers if a.questionid == question), None)
                if answer:
                    cell = ws.cell(row=row, column=col)
                    if question.type in ['MCQ', 'MSQ']:
                        options = [opt.optiontext for opt in answer.answer_option.all()]
                        value = ", ".join(options)
                        cell.value = value
                        # Update max width for this column
                        max_column_widths[col] = max(max_column_widths[col], len(value))
                    elif question.type in ['Text_One', 'Text_multi', 'DATE']:
                        cell.value = answer.answer_text
                        # Update max width for this column
                        if answer.answer_text:
                            max_column_widths[col] = max(max_column_widths[col], len(answer.answer_text))
                    elif question.type == 'Bottle':
                        cell.value = answer.bottle_value
                        # Update max width for this column
                        if answer.bottle_value is not None:
                            max_column_widths[col] = max(max_column_widths[col], len(str(answer.bottle_value)))
                    elif question.type in ['IMAGE', 'Sign']:
                        if answer.answer_image:
                            # Fetch the image file
                            image_path = os.path.join(settings.MEDIA_ROOT, answer.answer_image)
                            if os.path.exists(image_path):
                                # Load the image using Pillow
                                img = PILImage.open(image_path)

                                # Convert to a format compatible with openpyxl
                                img_byte_arr = io.BytesIO()
                                img.save(img_byte_arr, format=img.format if img.format else 'JPEG')
                                img_byte_arr.seek(0)

                                # Add image to Excel
                                openpyxl_img = OpenpyxlImage(img_byte_arr)
                                openpyxl_img.anchor = ws.cell(row=row, column=col).coordinate
                                openpyxl_img.width = 100  # Fixed width for images
                                openpyxl_img.height = 100  # Fixed height for images
                                ws.add_image(openpyxl_img)

                                # Set row height for image
                                max_row_heights[row] = max(max_row_heights.get(row, 0), 100)
                                # Set column width for image
                                max_column_widths[col] = max(max_column_widths[col], 20)  # Ensure enough width for image
                            else:
                                cell.value = "Image not found"
                                max_column_widths[col] = max(max_column_widths[col], len("Image not found"))
                else:
                    # Empty cell, ensure minimum width
                    max_column_widths[col] = max(max_column_widths[col], 10)  # Minimum width for empty cells

                # Center-align all cells
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            row += 1

        # Apply calculated widths and heights
        for col, width in max_column_widths.items():
            # Add some padding to the width (multiply by a factor for readability)
            adjusted_width = min(width * 1.2, 50)  # Cap at 50 to avoid overly wide columns
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(adjusted_width, 10)  # Minimum width of 10

        for row, height in max_row_heights.items():
            ws.row_dimensions[row].height = height

        # Save the workbook to a BytesIO buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create the response
        response = HttpResponse(
            content=output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="slambook_{slambook_title}_responses.xlsx"'
        return response

    except Url_slambook.DoesNotExist:
        return Response(
            {"error": "Invalid slambook ID"},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {"error": f"Failed to export responses: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )